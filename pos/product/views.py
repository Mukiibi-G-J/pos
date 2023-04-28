from product.form import (
    CategoryForm,
    AddProductForm,
    AddSalesForm,
)
from authentication.models import CustomUser
from django.shortcuts import render, redirect
from .utils import generate_random_code
from product.models import *
from django.http import JsonResponse, HttpResponse
from .form import FileUploadForm
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
import json
from django.db.models import F, Sum

from django.template.loader import render_to_string
from django.views import generic
from xhtml2pdf import pisa
from io import BytesIO
import re

import xlwt


# ???  decorator
from django.contrib.auth.decorators import login_required, user_passes_test


# openpyxl
from openpyxl import load_workbook
import uuid
from barcode import Code128, Code39
from barcode.writer import ImageWriter
import uuid
from django.conf import settings
import os


def is_admin(user):
    print(user)
    return user.groups.filter(name="ADMIN").exists()


@login_required
# @user_passes_test(is_admin)
def dashboard(request):
    # get the value of all your store products at cost price

    total_cost = Products.objects.aggregate(
        total_cost=Sum(F("unit_price") * F("quantity_in_stock"))
    )["total_cost"]
    # formatted_total_cost = '{:,.0f}'.format(total_cost)
    total_cost_at_selling_price = Products.objects.aggregate(
        total_cost=Sum(F("cost") * F("quantity_in_stock"))
    )["total_cost"]
    total_cost_of_sales = Sales.objects.aggregate(total_cost=Sum(F("total")))[
        "total_cost"
    ]

    total_no_of_products = Products.objects.all().count()
    total_no_of_sales = Sales.objects.all().count()
    sales_of_yesterday = Sales.objects.filter(
        date_sold__date=datetime.date.today() - datetime.timedelta(days=1)
    ).aggregate(total_sales=Sum(F("total")))["total_sales"]
    # value_if_true if condition else value_if_false
    total_sales_of_yesterday = sales_of_yesterday if sales_of_yesterday else 0
    if total_cost_of_sales is not None:
        formatted_total_cost_of_sales = format(int(total_cost_of_sales), ",d")
    else:
        formatted_total_cost_of_sales = 0

    if total_cost is not None:
        formatted_total_cost = format(int(total_cost), ",d")
    else:
        formatted_total_cost = 0

    if total_cost_at_selling_price is not None:
        foramtted_total_cost_at_selling_price = format(
            int(total_cost_at_selling_price), ",d"
        )
    else:
        foramtted_total_cost_at_selling_price = 0

    context = {
        "total_cost": formatted_total_cost,
        "total_cost_at_selling_price": foramtted_total_cost_at_selling_price,
        "total_sales_of_yesterday": format(int(total_sales_of_yesterday), ",d"),
        "total_no_of_products": total_no_of_products,
        "total_no_of_sales": total_no_of_sales,
        "total_cost_of_sales": formatted_total_cost_of_sales,
    }
    return render(request, "dashboard/dashboard.html", context)


# ?------------------------------------ADD PRODUCT ----------------------------
class AddProduct(generic.TemplateView):
    template_name = "products/add_products.html"


class ProductList(generic.View):
    template_name = "products/products_list.html"
    model = Products
    context_object_name = "products"
    ordering = ["created_at"]

    #  product form
    def get(self, request, *args, **kwargs):
        products = self.model.objects.all()
        products_form = AddProductForm()
        context = {"products": products, "products_form": products_form}
        return render(request, self.template_name, context)


@login_required
def addproduct(request):
    form = AddProductForm()
    uploadform = FileUploadForm()
    context = {"form": form, "uploadform": uploadform}

    return render(request, "products/add_products.html", context)


@login_required
def list_category(request):
    categories = Category.objects.all()
    context = {"categories": categories}

    return render(request, "categories/list_cat.html", context)


@login_required
def AddCategory(request):
    error = None
    if request.method == "POST":
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        else:
            error = form.errors
    else:
        form = CategoryForm()
    context = {"form": form, "error": error}
    return render(request, "categories/add_cat.html", context)


def add_single_product(request):
    if request.method == "POST":
        print(request.POST, request.FILES)
        product_name = request.POST["product_name"]
        category = request.POST["category_id"]
        cost_price = request.POST["unit_price"]
        selling_price = request.POST["cost"]
        quantity_in_stock = request.POST["quantity_in_stock"]
        brand_name = request.POST["brand"]
        unit_of_measure = request.POST["unit_of_measure"]
        reorder_level = request.POST["reorder_level"]
        description = request.POST["description"]
        # product_image = request.FILES["pic"]

        category = Category.objects.get(id=category)
        brand = Brand.objects.get(id=brand_name)
        # gererate product code

        product_code = generate_random_code(13)

        Products.objects.create(
            product_name=product_name,
            category_id=category,
            cost=selling_price,
            unit_price=cost_price,
            quantity_in_stock=quantity_in_stock,
            brand=brand,
            unit_of_measure=unit_of_measure,
            reorder_level=reorder_level,
            description=description,
            product_code=product_code,
            # product_image=product_image,
        )
        return redirect("products:product_list")

    return render(request, "products/add_products.html")


# ? ------------------------------------- sales -------------------------------------


def add_single_sell(request):
    form = AddSalesForm
    if request.method == "POST":
        user = request.user
        data = request.POST
        data_sold = data["date_sold"]
        product_name = data["product_name"]
        product_uuid = data["product_uuid"]
        discount = data["discount"]
        quantity = int(data["quantity"])
        price = int(data["price"])
        product = Products.objects.get(product_code=product_uuid)
        total_price = (int(price) * int(quantity)) - int(discount)
        print(total_price, discount)
        Sales.objects.create(
            user=user,
            date_sold=data_sold,
            product=product,
            quantity=quantity,
            total=total_price,
            price=price,
            discount=discount,
        )
        return redirect("products:list_sales")

    return render(
        request,
        "sales/add_single_sell.html",
        {"form": form},
    )


class AddSale(generic.View):
    template_name = "sales/add_sales.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        #! if request.is_ajax and request.method == "POST":
        #! checkin if request is ajax
        # if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
        name = request.POST.get("product")
        res = None

        qs = Products.objects.filter(product_name__icontains=name)
        data = []
        print(qs)
        if len(qs) > 0 and len(name) > 0:
            for i in qs:
                item = {
                    "code": i.product_code,
                    "name": i.product_name,
                    "price": i.cost,
                    "description": i.description,
                }
                data.append(item)
            res = data
        else:
            res = "No games found ........"
            return JsonResponse({"data": "Not found"}, status=400)
        return JsonResponse({"data": res}, status=200)


def complete_sale(request):
    items_sold = []
    if request.method == "POST":
        print(request.POST["cart"])
        cart = request.POST["cart"]
        # convert to json
        cart_data = json.loads(cart)

        # convert to dictionary
        # cart = eval(cart)
        # loop through a dictionary
        n = 0

        for key, value in cart_data.items():
            n += 1
            print(n)
            quantity = int(value["quantity"])
            product_uuid = value["product_uuid"]
            total_price = (int(value["sales_price"])) * int(quantity) - int(
                value["discount"]
            )
            discount = value["discount"]
            price = int(value["sales_price"])
            user = CustomUser.objects.get(username=request.user.username)
            print(quantity)
            # get product by uuid and quantity
            product = Products.objects.get(product_code=product_uuid)
            # check if product quantity is less than quantity
            if int(product.quantity_in_stock) < int(quantity):
                print("Not enough stock")
                message = "Not enough stock"

                return JsonResponse({"data": "Not enough stock"}, status=400)
            print(user, product)
            # create a sale
            Sales.objects.create(
                user=user,
                product=product,
                quantity=quantity,
                total=total_price,
                discount=discount,
                price=price,
            )
            # create the receipt data
            items_sold.append(
                {
                    "product_name": product,
                    "quantity": quantity,
                    "total_price": total_price,
                    "price": price,
                    "cashiers_name": user.full_name,
                },
            )
    elif request.method == "GET":
        items_dict = [item for item in items_sold]
        return render(request, "sales/receipt.html", {"items": items_dict})

    return JsonResponse({"data": "success"}, status=200)

    # Generate the PDF file
    # response = HttpResponse(content_type="application/pdf")
    # response["Content-Disposition"] = 'filename="receipt.pdf"'

    # # Generate the PDF content using the HTML template

    # pdf = BytesIO()
    # pisa.CreatePDF(BytesIO(html.encode("UTF-8")), pdf)
    # response.write(pdf.getvalue())

    # return response

    #         redirect("products:list_sales")

    # return JsonResponse({"data": "success"}, status=200)


class ListSale(generic.View):
    def get(self, request, *args, **kwargs):
        sales = Sales.objects.all()
        todays_total_sales = Sales.objects.filter(timestamp__date=datetime.date.today())
        todays_total_sales = sum([sale.total for sale in todays_total_sales])
        todays_total_sales = "{:,.0f}".format(todays_total_sales)

        context = {"sales": sales, "todays_total_sales": todays_total_sales}
        return render(request, "sales/list_sales.html", context)


# ? --------------------------- Purchases ---------------------------
class AddPurchase(generic.TemplateView):
    template_name = "purchases/add_purchases.html"


class ListPurchase(generic.TemplateView):
    template_name = "purchases/list_purchases.html"


def generate_purchase_list(request):
    product = Products.objects.all().order_by("reorder_level")

    context = {"products": product}
    return render(request, "purchases/generate_purchase_list.html", context)


@login_required
def upload_products(request):
    if request.method == "POST":
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]
            # sheet
            wb = load_workbook(file)
            # get all sheet names
            sheets = wb.get_sheet_names()
            context = {"products": []}
            for sheet in sheets:
                # get sheet by name
                ws = wb.get_sheet_by_name(sheet)
                values = [row[1:] for row in ws.values]
                # convert to dataframe

                df = pd.DataFrame(values)

                header_list = [
                    "PRODUCT NAME",
                    "DESCRIPTION",
                    "QUANTITY",
                    "PURCHASE PRICE",
                    "SALES PRICE",
                    "CATEGORY",
                    "UNIT ID",
                    "BRAND",
                ]
                # check if header is in the first row

                if df.iloc[0].tolist() == header_list:
                    df = df[0:]
                    # get the first row as header
                    df.columns = df.iloc[0]
                    # drop the first row
                    df = df[1:]
                    # reset index
                    df = df.reset_index(drop=True)
                    # drop nonwe in PURCHASE PRICE, SALES PRICE, QUANTITY, PRODUCT NAME
                    df = df.dropna(
                        subset=[
                            "PURCHASE PRICE",
                            "SALES PRICE",
                            "QUANTITY",
                            "PRODUCT NAME",
                            "CATEGORY",
                            "UNIT ID",
                        ]
                    )
                    # get the first row as header
                    product_name = df["PRODUCT NAME"].tolist()
                    description = df["DESCRIPTION"].tolist()
                    quantity = df["QUANTITY"].tolist()
                    purchase_price = df["PURCHASE PRICE"].tolist()
                    sales_price = df["SALES PRICE"].tolist()
                    category = df["CATEGORY"].tolist()
                    unit = df["UNIT ID"].tolist()
                    brand = df["BRAND"].tolist()

                    for i in range(len(product_name)):
                        # create uuid
                        code = generate_random_code(13)
                        print(code)

                        # create context

                        products = {
                            "uuid": code,
                            "product_name": product_name[i],
                            "description": description[i],
                            "quantity": quantity[i],
                            "purchase_price": purchase_price[i],
                            "sales_price": sales_price[i],
                            "category": category[i],
                            "unit": unit[i],
                            "brand": brand[i],
                        }

                        context["products"].append(products)
                    return render(request, "products/add_products.html", context)

                else:
                    print("header not found")
    return render(request, "products/add_products.html")


def generate_barcode(request):
    product = Products.objects.all()
    for uuid in product:
        code = uuid.product_code
        product_name = uuid.product_name.replace(" ", "_").replace("*", "_")
        barcode = Code128(code, writer=ImageWriter())

        # Set the width of the barcode image to 600 pixels
        barcode.writer.set_options({"width": 600})

        # Save the barcode image to a file

        # save the barcode to a file in the images directory
        image_dir = os.path.join(settings.BASE_DIR, "images")
        if not os.path.exists(image_dir):
            os.makedirs(image_dir)
        # change image name to product name
        barcode.save(
            os.path.join(image_dir, product_name),
            options={"background": "white", "foreground": "black"},
        )
    return JsonResponse({"success": True})


def add_product_upload(request):
    if request.method == "POST":
        data = request.POST
        # get values form the dict
        for i in request.POST.getlist("product_name"):
            product_name = request.POST.getlist("product_name")
            description = request.POST.getlist("description")
            quantity = request.POST.getlist("quantity")
            purchase_price = request.POST.getlist("purchase_price")
            sales_price = request.POST.getlist("sales_price")
            category = request.POST.getlist("category")
            unit = request.POST.getlist("unit")
            brand = request.POST.getlist("brand")
            uuid = request.POST.getlist("uuid")
            # print(product_name, description, quantity, purchase_price, sales_price, category, unit, brand, uuid)
        for (
            product_name,
            description,
            quantity,
            purchase_price,
            sales_price,
            category,
            unit,
            brand,
            uuid,
        ) in zip(
            product_name,
            description,
            quantity,
            purchase_price,
            sales_price,
            category,
            unit,
            brand,
            uuid,
        ):
            print(
                product_name,
                description,
                quantity,
                purchase_price,
                sales_price,
                category,
                unit,
                brand,
                uuid,
            )
            # delte all prod
            # Products.objects.all().delete()
            # Category.objects.all().delete()
            # Brand.objects.all().delete()
            if Category.objects.filter(category_name=category).exists() == False:
                category = Category.objects.create(category_name=category)
                category.save()
            else:
                category = Category.objects.get(category_name=category)

            if Brand.objects.filter(brand_name=brand).exists() == False:
                brand = Brand.objects.create(brand_name=brand)
                brand.save()
            else:
                brand = Brand.objects.get(brand_name=brand)
                brand.save()

            products = Products.objects.create(
                product_name=product_name,
                description=description,
                quantity_in_stock=quantity,
                unit_price=purchase_price,
                cost=sales_price,
                category_id=category,
                unit_of_measure="pcs",
                brand=brand,
                product_code=uuid,
            )
            products.save()

        return JsonResponse({"success": True})
    return JsonResponse({"success": False})


def get_product_by_uuid(request, uuid):
    if request.method == "GET":
        product = Products.objects.get(product_code=uuid)

        data = {
            "product_name": product.product_name,
            "sales_price": product.cost,
            "product_uuid": product.product_code,
        }
        print(data)
        return JsonResponse(data)


@login_required
def products_list(request):
    products = Products.objects.all()
    context = {"products": products}
    return render(request, "products/products_list.html", context)


# ?------------------REPORTS----------------------------------------


def reports(request):
    return render(request, "reports/reports.html")


# ?  ----------export to excel----------------


def export_to_excel(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="products.xls"'
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Products")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = [
        "Product Code",
        "Product Name",
        "Description",
        "Quantity",
        "Cost Price",
        "Sales Price",
        "Reorder Level",
        "Supplier",
    ]
    
    
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
    
    
    font_style = xlwt.XFStyle()
    rows = Products.objects.all().values_list(
        "product_code","product_name","description","quantity_in_stock","unit_price","cost","reorder_level"
    )
    
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response